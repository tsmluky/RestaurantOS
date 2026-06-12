"""Excel / CSV schedule import endpoint.

POST /shifts/import
    - Accepts a multipart file upload (.xlsx or .csv)
    - Parses rows, matches employees by email / full_name / employee_code
    - Creates valid rows as DRAFT shifts in the DB
    - Returns { imported, skipped, errors, shifts }

Expected columns (case-insensitive, Spanish or English):
    employee  | email / empleado / nombre / name / employee
    date      | fecha / date / dia
    start     | inicio / entrada / start / hora_inicio / from
    end       | fin / salida / end / hora_fin / to
    role      | puesto / role / rol / cargo   (optional)
    notes     | notas / notes                 (optional)

Date formats accepted: YYYY-MM-DD, DD/MM/YYYY, D/M/YY
Time formats accepted: HH:MM, H:MM
"""
from __future__ import annotations

import csv
import io
import logging
from datetime import UTC, datetime, time
from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy import and_, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import TenantId, require_manager
from app.models.employee import EmployeeProfile
from app.models.enums import ShiftStatus
from app.models.restaurant import Restaurant
from app.models.shift import Shift
from app.models.user import User
from app.schemas.shift import ShiftResponse

logger = logging.getLogger("restaurantos.import")
router = APIRouter()

# ── Column name aliases ───────────────────────────────────────────────────────

_COL_EMPLOYEE = {"email", "empleado", "nombre", "name", "employee", "trabajador"}
_COL_DATE     = {"fecha", "date", "dia", "día", "day"}
_COL_START    = {"inicio", "entrada", "start", "hora_inicio", "from", "hora inicio", "inicio jornada"}
_COL_END      = {"fin", "salida", "end", "hora_fin", "to", "hora fin", "fin jornada"}
_COL_ROLE     = {"puesto", "role", "rol", "cargo", "position"}
_COL_NOTES    = {"notas", "notes", "nota", "note", "observaciones"}


def _match_col(header: str, aliases: set[str]) -> bool:
    return header.strip().lower() in aliases


def _parse_date(raw: str) -> datetime | None:
    raw = raw.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _parse_time(raw: str) -> time | None:
    raw = raw.strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%-H:%M"):
        try:
            return datetime.strptime(raw, fmt).time()
        except ValueError:
            continue
    return None


def _combine(date_obj: datetime, t: time, tz=UTC) -> datetime:
    return datetime(
        date_obj.year, date_obj.month, date_obj.day,
        t.hour, t.minute, 0, tzinfo=tz
    )


# ── Schemas ───────────────────────────────────────────────────────────────────

class ImportError(BaseModel):
    row: int
    employee: str
    message: str


class ImportResult(BaseModel):
    imported: int
    skipped: int
    errors: list[ImportError]
    shifts: list[ShiftResponse]


# ── Parsing helpers ───────────────────────────────────────────────────────────

def _parse_rows_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return [h.strip() for h in rows[0]], rows[1:]


def _parse_rows_from_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status.HTTP_501_NOT_IMPLEMENTED,
            detail="openpyxl no disponible en el servidor",
        ) from exc

    wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [
        [str(cell).strip() if cell is not None else "" for cell in row]
        for row in rows[1:]
    ]
    return headers, data


# ── Main endpoint ─────────────────────────────────────────────────────────────

def _to_shift_response(shift: Shift) -> ShiftResponse:
    duration = int((shift.ends_at - shift.starts_at).total_seconds() // 60)
    return ShiftResponse(
        id=shift.id,
        tenant_id=shift.tenant_id,
        restaurant_id=shift.restaurant_id,
        restaurant_name=None,
        user_id=shift.user_id,
        user_full_name=None,
        starts_at=shift.starts_at,
        ends_at=shift.ends_at,
        duration_minutes=duration,
        role=shift.role,
        notes=shift.notes,
        status=shift.status,
        created_at=shift.created_at,
    )


@router.post(
    "/import",
    response_model=ImportResult,
    dependencies=[Depends(require_manager)],
    summary="Import weekly schedule from Excel (.xlsx) or CSV",
)
def import_shifts(
    file: UploadFile,
    tenant_id: TenantId,
    db: Annotated[Session, Depends(get_db)],
    restaurant_id: UUID = Query(..., description="Target restaurant for all imported shifts"),
) -> ImportResult:
    # Validate restaurant
    restaurant = db.scalar(
        select(Restaurant).where(
            Restaurant.id == restaurant_id,
            Restaurant.tenant_id == tenant_id,
            Restaurant.deleted_at.is_(None),
        )
    )
    if restaurant is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Sucursal no encontrada")

    # Read file
    content = file.file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx") or file.content_type in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        headers, data_rows = _parse_rows_from_xlsx(content)
    elif filename.endswith(".csv") or "csv" in (file.content_type or ""):
        headers, data_rows = _parse_rows_from_csv(content)
    else:
        # Try xlsx first, then csv
        try:
            headers, data_rows = _parse_rows_from_xlsx(content)
        except Exception:
            headers, data_rows = _parse_rows_from_csv(content)

    if not headers:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="El archivo está vacío")

    # Map column indices
    col: dict[str, int | None] = {
        "employee": None, "date": None, "start": None, "end": None,
        "role": None, "notes": None,
    }
    for i, h in enumerate(headers):
        if _match_col(h, _COL_EMPLOYEE): col["employee"] = i
        elif _match_col(h, _COL_DATE):   col["date"] = i
        elif _match_col(h, _COL_START):  col["start"] = i
        elif _match_col(h, _COL_END):    col["end"] = i
        elif _match_col(h, _COL_ROLE):   col["role"] = i
        elif _match_col(h, _COL_NOTES):  col["notes"] = i

    missing = [k for k in ("employee", "date", "start", "end") if col[k] is None]
    if missing:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Columnas requeridas no encontradas: {', '.join(missing)}. "
                f"Cabeceras detectadas: {', '.join(headers)}"
            ),
        )

    # Build employee lookup: email → user, full_name lower → user, code → user
    all_employees: list[User] = list(
        db.scalars(
            select(User)
            .join(EmployeeProfile, User.id == EmployeeProfile.user_id)
            .where(User.tenant_id == tenant_id, User.deleted_at.is_(None))
        ).all()
    )
    profiles = {
        p.user_id: p
        for p in db.scalars(
            select(EmployeeProfile).where(EmployeeProfile.tenant_id == tenant_id)
        ).all()
    }
    by_email: dict[str, User] = {(u.email or "").lower(): u for u in all_employees if u.email}
    by_name:  dict[str, User] = {u.full_name.lower(): u for u in all_employees}
    by_code:  dict[str, User] = {}
    for p in profiles.values():
        if p.employee_code:
            owner = next((u for u in all_employees if u.id == p.user_id), None)
            if owner:
                by_code[p.employee_code.lower()] = owner

    imported = 0
    skipped = 0
    errors: list[ImportError] = []
    created_shifts: list[Shift] = []

    for row_idx, row in enumerate(data_rows, start=2):
        if not row or all(c == "" for c in row):
            skipped += 1
            continue

        def cell(key: str) -> str:
            idx = col[key]
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip()

        emp_raw = cell("employee")
        date_raw = cell("date")
        start_raw = cell("start")
        end_raw = cell("end")
        role_val = cell("role") or None
        notes_val = cell("notes") or None

        if not emp_raw:
            skipped += 1
            continue

        # Resolve employee
        employee = (
            by_email.get(emp_raw.lower())
            or by_name.get(emp_raw.lower())
            or by_code.get(emp_raw.lower())
        )
        if employee is None:
            errors.append(ImportError(
                row=row_idx,
                employee=emp_raw,
                message=f"Empleado '{emp_raw}' no encontrado en el sistema",
            ))
            continue

        date_obj = _parse_date(date_raw)
        if date_obj is None:
            errors.append(ImportError(
                row=row_idx, employee=emp_raw,
                message=f"Fecha '{date_raw}' no reconocida (usa DD/MM/AAAA)",
            ))
            continue

        start_t = _parse_time(start_raw)
        if start_t is None:
            errors.append(ImportError(
                row=row_idx, employee=emp_raw,
                message=f"Hora inicio '{start_raw}' no reconocida (usa HH:MM)",
            ))
            continue

        end_t = _parse_time(end_raw)
        if end_t is None:
            errors.append(ImportError(
                row=row_idx, employee=emp_raw,
                message=f"Hora fin '{end_raw}' no reconocida (usa HH:MM)",
            ))
            continue

        starts_at = _combine(date_obj, start_t)
        ends_at   = _combine(date_obj, end_t)

        if ends_at <= starts_at:
            errors.append(ImportError(
                row=row_idx, employee=emp_raw,
                message="La hora de fin debe ser posterior a la de inicio",
            ))
            continue

        # Check overlap
        overlap = db.scalar(
            select(Shift).where(
                Shift.tenant_id == tenant_id,
                Shift.user_id == employee.id,
                Shift.status.in_(["DRAFT", "PUBLISHED", "SCHEDULED"]),
                and_(Shift.starts_at < ends_at, Shift.ends_at > starts_at),
            )
        )
        if overlap is not None:
            errors.append(ImportError(
                row=row_idx, employee=emp_raw,
                message=f"Solapamiento con turno existente ({starts_at.strftime('%d/%m %H:%M')}–{ends_at.strftime('%H:%M')})",
            ))
            continue

        shift = Shift(
            tenant_id=tenant_id,
            restaurant_id=restaurant_id,
            user_id=employee.id,
            starts_at=starts_at,
            ends_at=ends_at,
            role=role_val,
            notes=notes_val,
            status=ShiftStatus.DRAFT.value,
        )
        db.add(shift)
        created_shifts.append(shift)
        imported += 1

    db.commit()
    for s in created_shifts:
        db.refresh(s)

    logger.info(
        "Import complete: %d imported, %d skipped, %d errors (restaurant=%s)",
        imported, skipped, len(errors), restaurant_id,
    )

    return ImportResult(
        imported=imported,
        skipped=skipped,
        errors=errors,
        shifts=[_to_shift_response(s) for s in created_shifts],
    )
